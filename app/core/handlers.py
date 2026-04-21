import os
import pandas as pd
from openpyxl import load_workbook
import tempfile
import uuid
import shutil

# Try to import win32com and pythoncom for Windows only functionality
try:
    import win32com.client
    import pythoncom
    WIN32_AVAILABLE = True
except ImportError:
    WIN32_AVAILABLE = False
except Exception:
    WIN32_AVAILABLE = False

from app.utils.common import (
    clean_text, detect_header_index_from_rows, normalize_columns_from_header_row,
    choose_header_index, dataframe_from_rows_with_header, trim_empty_columns_df,
    is_html_content, get_readable_file_path, _trim_rows_to_used_content,
    _normalize_excel_value_matrix
)

def read_csv_header_fast(file_path, scan_rows=200):
    encodings = ["utf-8-sig", "cp949", "euc-kr", "utf-8", "latin1"]
    last_error = None
    for enc in encodings:
        try:
            sample = pd.read_csv(file_path, nrows=scan_rows, header=None, dtype=str, encoding=enc)
            rows = sample.values.tolist()
            if not rows:
                return [], 0, enc
            header_idx = detect_header_index_from_rows(rows)
            header = normalize_columns_from_header_row(rows[header_idx])
            return header, header_idx, enc
        except Exception as e:
            last_error = e
    raise last_error

def read_xlsx_header_fast(file_path, sheet_name=None, scan_rows=200):
    wb = load_workbook(file_path, read_only=True, data_only=True)
    if sheet_name:
        if sheet_name not in wb.sheetnames:
            wb.close()
            raise ValueError(f"시트 '{sheet_name}' 이(가) 파일에 없습니다: {os.path.basename(file_path)}")
        ws = wb[sheet_name]
    else:
        ws = wb[wb.sheetnames[0]]
    rows = []
    for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
        rows.append(list(row))
        if i >= scan_rows:
            break
    wb.close()
    if not rows:
        return [], 0
    
    rows = _trim_rows_to_used_content(rows)
    header_idx = detect_header_index_from_rows(rows)
    header = normalize_columns_from_header_row(rows[header_idx])
    return header, header_idx

def get_sheet_names(file_path):
    name = file_path if isinstance(file_path, str) else getattr(file_path, "name", "")
    ext = os.path.splitext(name)[1].lower()
    if ext in [".xlsx", ".xlsm"] and not is_html_content(file_path):
        try:
            wb = load_workbook(file_path, read_only=True, data_only=True)
            names = wb.sheetnames
            wb.close()
            return names
        except Exception:
            return []
    return []

def extract_columns_fast(file_path, sheet_name=None, header_row_idx=None, force_html=False):
    name = file_path if isinstance(file_path, str) else getattr(file_path, "name", "")
    ext = os.path.splitext(name)[1].lower()

    if ext == ".csv" and not force_html:
        header, _, _ = read_csv_header_fast(file_path, scan_rows=200)
        if header_row_idx is None:
            return header
        _, _, enc = read_csv_header_fast(file_path, scan_rows=200)
        df = pd.read_csv(file_path, nrows=200, header=None, dtype=str, encoding=enc)
        rows = df.values.tolist()
        if not rows:
            return []
        header_idx = choose_header_index(rows, header_row_idx=header_row_idx, max_scan=200)
        return normalize_columns_from_header_row(rows[header_idx])

    if force_html or ext in [".html", ".htm"] or (ext == ".xls" and is_html_content(file_path)):
        tables = pd.read_html(file_path, header=None)
        if not tables:
            return []
        df = pd.concat(tables, ignore_index=True, sort=False)
        rows = df.values.tolist()
        header_idx = choose_header_index(rows[:200] if len(rows) > 200 else rows, header_row_idx=header_row_idx, max_scan=200)
        return normalize_columns_from_header_row(rows[header_idx])

    if ext in [".xlsx", ".xlsm"]:
        if header_row_idx is None:
            header, _ = read_xlsx_header_fast(file_path, sheet_name=sheet_name, scan_rows=200)
            return header
        wb = load_workbook(file_path, read_only=True, data_only=True)
        ws = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb[wb.sheetnames[0]]
        rows = []
        for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
            rows.append(list(row))
            if i >= 200:
                break
        wb.close()
        if not rows:
            return []
        rows = _trim_rows_to_used_content(rows)
        header_idx = choose_header_index(rows, header_row_idx=header_row_idx, max_scan=200)
        return normalize_columns_from_header_row(rows[header_idx])

    if ext == ".xls":
        try:
            tables = pd.read_html(file_path, header=None)
            if tables:
                df = pd.concat(tables, ignore_index=True, sort=False)
                rows = df.values.tolist()
                header_idx = choose_header_index(rows[:200] if len(rows) > 200 else rows, header_row_idx=header_row_idx, max_scan=200)
                return normalize_columns_from_header_row(rows[header_idx])
        except Exception:
            pass

    return []

def load_file_to_df(file_path, sheet_name=None, header_row_idx=None, force_html=False):
    name = file_path if isinstance(file_path, str) else getattr(file_path, "name", "")
    ext = os.path.splitext(name)[1].lower()

    if ext == ".csv" and not force_html:
        _, _, enc = read_csv_header_fast(file_path, scan_rows=200)
        df = pd.read_csv(file_path, header=None, dtype=str, encoding=enc)
        rows = df.values.tolist()
        if not rows:
            return pd.DataFrame()
        header_idx = choose_header_index(rows[:200] if len(rows) > 200 else rows, header_row_idx=header_row_idx, max_scan=200)
        header = normalize_columns_from_header_row(rows[header_idx])
        body = df.iloc[header_idx + 1:].reset_index(drop=True).copy()
        if body.shape[1] < len(header):
            for _ in range(len(header) - body.shape[1]):
                body[body.shape[1]] = None
        elif body.shape[1] > len(header):
            body = body.iloc[:, :len(header)]
        body.columns = header
        return trim_empty_columns_df(body)

    if force_html or ext in [".html", ".htm"] or (ext == ".xls" and is_html_content(file_path)):
        try:
            tables = pd.read_html(file_path, header=None, engine='lxml')
        except:
            try:
                tables = pd.read_html(file_path, header=None)
            except:
                tables = []
        if not tables:
            return pd.DataFrame()
        df = pd.concat(tables, ignore_index=True, sort=False)
        rows = df.values.tolist()
        if not rows:
            return pd.DataFrame()
        header_idx = choose_header_index(rows[:200] if len(rows) > 200 else rows, header_row_idx=header_row_idx, max_scan=200)
        header = normalize_columns_from_header_row(rows[header_idx])
        body = df.iloc[header_idx + 1:].reset_index(drop=True).copy()
        if body.shape[1] < len(header):
            for i in range(len(header) - body.shape[1]):
                body[body.shape[1] + i] = None
        elif body.shape[1] > len(header):
            body = body.iloc[:, :len(header)]
        body.columns = header
        return trim_empty_columns_df(body)

    if ext in [".xlsx", ".xlsm"]:
        wb = load_workbook(file_path, read_only=True, data_only=True)
        if sheet_name:
            if sheet_name not in wb.sheetnames:
                wb.close()
                raise ValueError(f"시트 '{sheet_name}' 이(가) 파일에 없습니다: {os.path.basename(file_path)}")
            ws = wb[sheet_name]
        else:
            ws = wb[wb.sheetnames[0]]

        rows = []
        for row in ws.iter_rows(values_only=True):
            rows.append(list(row))
        wb.close()
        return dataframe_from_rows_with_header(rows, header_row_idx=header_row_idx, scan_rows=200)

    if ext == ".xls":
        try:
            tables = pd.read_html(file_path, header=None)
            if tables:
                df = pd.concat(tables, ignore_index=True, sort=False)
                rows = df.values.tolist()
                if not rows:
                    return pd.DataFrame()
                header_idx = choose_header_index(rows[:200] if len(rows) > 200 else rows, header_row_idx=header_row_idx, max_scan=200)
                header = normalize_columns_from_header_row(rows[header_idx])
                body = df.iloc[header_idx + 1:].reset_index(drop=True).copy()
                if body.shape[1] < len(header):
                    for i in range(len(header) - body.shape[1]):
                        body[body.shape[1] + i] = None
                elif body.shape[1] > len(header):
                    body = body.iloc[:, :len(header)]
                body.columns = header
                return trim_empty_columns_df(body)
        except Exception:
            pass

    raise ValueError(f"지원하지 않는 형식: {os.path.basename(file_path)}")

# Windows-only Excel interaction functions
def list_open_excel_workbooks():
    if not WIN32_AVAILABLE:
        return []
    try:
        app = win32com.client.GetActiveObject("Excel.Application")
        items = []
        for wb in app.Workbooks:
            try:
                full_name = wb.FullName
            except Exception:
                full_name = ""
            items.append({
                "name": clean_text(wb.Name),
                "full_name": clean_text(full_name),
                "saved": bool(getattr(wb, "Saved", False)),
                "sheet_names": [clean_text(ws.Name) for ws in wb.Worksheets]
            })
        return items
    except Exception:
        return []

def read_open_excel_sheet_rows(workbook_name, sheet_name, max_rows=None):
    if not WIN32_AVAILABLE:
        raise RuntimeError("pywin32가 설치되지 않았습니다.")

    pythoncom.CoInitialize()
    try:
        app = win32com.client.GetActiveObject("Excel.Application")
        target_wb = None
        for wb in app.Workbooks:
            if clean_text(wb.Name) == clean_text(workbook_name):
                target_wb = wb
                break
        if target_wb is None:
            raise RuntimeError("열려있는 워크북을 찾지 못했습니다.")
        ws = target_wb.Worksheets(sheet_name)
        
        if max_rows:
            # Optimize: only read up to max_rows
            values = ws.Range(ws.Cells(1, 1), ws.Cells(max_rows, ws.UsedRange.Columns.Count)).Value
        else:
            values = ws.UsedRange.Value
            
        rows = _normalize_excel_value_matrix(values)
        rows = _trim_rows_to_used_content(rows)
        return rows
    finally:
        pythoncom.CoUninitialize()
def load_file_sample_rows(file_path, sheet_name=None, force_html=False, max_rows=100):
    name = file_path if isinstance(file_path, str) else getattr(file_path, "name", "")
    ext = os.path.splitext(name)[1].lower()
    if ext == ".csv" and not force_html:
        try:
            _, _, enc = read_csv_header_fast(file_path)
            df = pd.read_csv(file_path, nrows=max_rows, header=None, dtype=str, encoding=enc)
            return df.values.tolist()
        except:
            return []
    if force_html or ext in [".html", ".htm"] or (ext == ".xls" and is_html_content(file_path)):
        try:
            tables = pd.read_html(file_path, header=None)
            if not tables: return []
            df = pd.concat(tables, ignore_index=True).head(max_rows)
            return df.values.tolist()
        except:
            return []
    if ext in [".xlsx", ".xlsm"]:
        try:
            wb = load_workbook(file_path, read_only=True, data_only=True)
            ws = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb[wb.sheetnames[0]]
            rows = []
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                rows.append(list(row))
                if i >= max_rows: break
            wb.close()
            return rows
        except:
            return []
    return []

def load_open_excel_sheet_df(workbook_name, sheet_name, max_rows=None, header_row_idx=None):
    rows = read_open_excel_sheet_rows(workbook_name, sheet_name, max_rows=max_rows)
    return dataframe_from_rows_with_header(rows, header_row_idx=header_row_idx)

def extract_open_excel_columns(workbook_name, sheet_name, header_row_idx=None):
    rows = read_open_excel_sheet_rows(workbook_name, sheet_name, max_rows=200)
    if not rows: return []
    header_idx = choose_header_index(rows, header_row_idx=header_row_idx)
    return normalize_columns_from_header_row(rows[header_idx])

def extract_open_excel_unique_values(workbook_name, sheet_name, col_name, header_row_idx=None, max_values=200):
    df = load_open_excel_sheet_df(workbook_name, sheet_name, max_rows=50000, header_row_idx=header_row_idx)
    if df.empty or col_name not in df.columns:
        return []
    vals = df[col_name].dropna().unique().tolist()
    return sorted([str(v) for v in vals if str(v).strip()])[:max_values]

def extract_unique_values_fast(file_path, col_name, sheet_name=None, header_row_idx=None, force_html=False, max_values=200, max_scan_rows=50000):
    df = load_file_to_df(file_path, sheet_name=sheet_name, header_row_idx=header_row_idx, force_html=force_html)
    if df.empty or col_name not in df.columns:
        return []
    vals = df[col_name].dropna().unique().tolist()
    return sorted([str(v) for v in vals if str(v).strip()])[:max_values]
