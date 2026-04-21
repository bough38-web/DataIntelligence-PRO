import os
import sys
import re
import glob
import json
import shutil
import tempfile
import traceback
import uuid
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

from PySide6.QtCore import Qt, QThread, Signal
from PySide6.QtGui import QFont
from PySide6.QtWidgets import (
    QApplication, QWidget, QVBoxLayout, QHBoxLayout, QPushButton, QLabel,
    QFileDialog, QListWidget, QListWidgetItem, QComboBox, QLineEdit,
    QTextEdit, QMessageBox, QProgressBar, QCheckBox, QSplitter, QFrame,
    QScrollArea, QGridLayout, QInputDialog, QTableWidget, QTableWidgetItem,
    QHeaderView, QTabWidget
)

try:
    import win32com.client  # type: ignore
    import pythoncom  # type: ignore
    WIN32_AVAILABLE = True
except Exception:
    WIN32_AVAILABLE = False

# =========================================================
# 공통 유틸
# =========================================================
def clean_text(x):
    if x is None:
        return ""
    try:
        if pd.isna(x):
            return ""
    except Exception:
        pass
    s = str(x)
    s = s.replace("\xa0", " ")
    s = re.sub(r"\s+", " ", s).strip()
    return s


def make_unique_columns(cols):
    seen = {}
    out = []
    for i, c in enumerate(cols):
        name = clean_text(c) or f"컬럼{i+1}"
        if name in seen:
            seen[name] += 1
            name = f"{name}_{seen[name]}"
        else:
            seen[name] = 0
        out.append(name)
    return out


def header_score(row):
    score = 0.0
    for v in row:
        t = clean_text(v)
        if not t:
            continue
        if re.fullmatch(r"\d+(\.\d+)?", t):
            score += 0.2
        else:
            score += 1.0
    return score


def detect_header_index_from_rows(rows, max_scan=30):
    best_idx = 0
    best_score = -1.0
    for i, row in enumerate(rows[:max_scan]):
        score = header_score(row)
        if score > best_score:
            best_score = score
            best_idx = i
    return best_idx


def normalize_columns_from_header_row(header_row):
    return make_unique_columns(header_row)

def choose_header_index(rows, header_row_idx=None, max_scan=30):
    """
    header_row_idx: None이면 자동 탐지, 0 이상이면 해당 행을 헤더로 사용
    """
    if not rows:
        return 0
    if header_row_idx is None:
        return detect_header_index_from_rows(rows, max_scan=max_scan)
    try:
        idx = int(header_row_idx)
    except Exception:
        idx = 0
    if idx < 0:
        idx = 0
    if idx >= len(rows):
        idx = len(rows) - 1
    return idx


def dataframe_from_rows_with_header(rows, header_row_idx=None, scan_rows=200):
    """
    rows(list[list])에서 헤더 행을 선택해 DataFrame 생성
    """
    if not rows:
        return pd.DataFrame()

    trimmed_rows = _trim_rows_to_used_content(rows)
    if not trimmed_rows:
        return pd.DataFrame()

    header_idx = choose_header_index(trimmed_rows, header_row_idx=header_row_idx, max_scan=scan_rows)
    header = normalize_columns_from_header_row(trimmed_rows[header_idx])
    body_rows = trimmed_rows[header_idx + 1:]

    normalized = []
    for r in body_rows:
        vals = list(r)
        if len(vals) < len(header):
            vals += [None] * (len(header) - len(vals))
        elif len(vals) > len(header):
            vals = vals[:len(header)]
        normalized.append(vals)

    body = pd.DataFrame(normalized, columns=header)
    return trim_empty_columns_df(body)


def preview_text(df, rows=10):
    if df is None or df.empty:
        return "(데이터 없음)"
    return df.head(rows).fillna("").to_string(index=False)


def raw_rows_preview_text(rows, max_rows=20):
    if not rows:
        return "(원본 행 데이터 없음)"
    clipped = rows[:max_rows]
    width = max((len(r) for r in clipped), default=0)
    normalized = []
    for r in clipped:
        vals = list(r)
        if len(vals) < width:
            vals += [None] * (width - len(vals))
        normalized.append(vals)
    df = pd.DataFrame(normalized)
    return df.fillna("").to_string(index=False, header=False)


def load_file_sample_rows(file_path, sheet_name=None, max_rows=20, force_html=False):
    ext = os.path.splitext(file_path)[1].lower()

    if ext == ".csv" and not force_html:
        _, _, enc = read_csv_header_fast(file_path, scan_rows=max(max_rows, 50))
        df = pd.read_csv(file_path, nrows=max_rows, header=None, dtype=str, encoding=enc)
        return df.values.tolist()

    if force_html or ext in [".html", ".htm"] or (ext == ".xls" and is_html_content(file_path)):
        tables = pd.read_html(file_path, header=None)
        if not tables:
            return []
        df = pd.concat(tables, ignore_index=True, sort=False)
        return df.head(max_rows).values.tolist()

    if ext in [".xlsx", ".xlsm"]:
        wb = load_workbook(file_path, read_only=True, data_only=True)
        ws = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb[wb.sheetnames[0]]
        rows = []
        for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
            rows.append(list(row))
            if i >= max_rows:
                break
        wb.close()
        return rows

    if ext == ".xls":
        try:
            tables = pd.read_html(file_path, header=None)
            if tables:
                df = pd.concat(tables, ignore_index=True, sort=False)
                return df.head(max_rows).values.tolist()
        except Exception:
            pass

    return []


def load_open_excel_raw_rows(workbook_name, sheet_name, max_rows=20):
    return read_open_excel_sheet_rows(workbook_name, sheet_name, max_rows=max_rows)


def trim_empty_columns_df(df):
    if df.empty:
        return df
    keep_cols = []
    for c in df.columns:
        s = (
            df[c]
            .fillna("")
            .astype(str)
            .str.replace("\xa0", " ", regex=False)
            .str.strip()
        )
        if s.ne("").any():
            keep_cols.append(c)
    return df[keep_cols]


def fill_service_small_from_mid(df):
    cols = list(df.columns)
    mid_col = next((c for c in cols if "서비스" in str(c) and "중" in str(c)), None)
    small_col = next((c for c in cols if "서비스" in str(c) and "소" in str(c)), None)
    if mid_col is None or small_col is None:
        return df
    mid_s = df[mid_col].astype(str).map(clean_text)
    small_s = df[small_col].astype(str).map(clean_text)
    blank_mask = small_s.eq("") | small_s.str.lower().eq("nan")
    df.loc[blank_mask, small_col] = mid_s[blank_mask]
    return df


def is_html_content(file_path):
    try:
        with open(file_path, "rb") as f:
            raw = f.read(65536).lower()
        markers = [
            b"<html", b"<table", b"<!doctype html", b"<head", b"<body",
            b"<meta", b"<tr", b"<td", b"<th"
        ]
        return any(m in raw for m in markers)
    except Exception:
        return False


def is_file_locked(file_path):
    try:
        with open(file_path, "ab"):
            return False
    except Exception:
        return True


def make_temp_copy(file_path):
    temp_dir = os.path.join(tempfile.gettempdir(), "gui_data_extractor_temp")
    os.makedirs(temp_dir, exist_ok=True)
    ext = os.path.splitext(file_path)[1].lower()
    temp_name = f"{uuid.uuid4().hex}{ext}"
    temp_path = os.path.join(temp_dir, temp_name)
    shutil.copy2(file_path, temp_path)
    return temp_path


def get_readable_file_path(file_path):
    if is_file_locked(file_path):
        return make_temp_copy(file_path)
    return file_path


def to_numeric_series(s):
    s2 = (
        s.astype(str)
        .map(clean_text)
        .str.replace(",", "", regex=False)
        .str.replace("원", "", regex=False)
        .str.replace("%", "", regex=False)
    )
    return pd.to_numeric(s2, errors="coerce")


def to_datetime_series(s):
    s2 = s.astype(str).map(clean_text)
    return pd.to_datetime(s2, errors="coerce")


class JsonStore:
    def __init__(self, filename, default=None):
        self.path = Path(filename)
        self.default = default if default is not None else []

    def load(self):
        try:
            if not self.path.exists():
                return json.loads(json.dumps(self.default, ensure_ascii=False))
            with open(self.path, "r", encoding="utf-8") as f:
                return json.load(f)
        except Exception:
            return json.loads(json.dumps(self.default, ensure_ascii=False))

    def save(self, data):
        with open(self.path, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2)


# =========================================================
# 필터 / 정렬 / 치환
# =========================================================
def apply_replacements(df, replacement_rules):
    if df.empty or not replacement_rules:
        return df
    result = df.copy()
    for rule in replacement_rules:
        col = clean_text(rule.get("column"))
        from_val = clean_text(rule.get("from"))
        to_val = clean_text(rule.get("to"))
        if not col or col not in result.columns or from_val == "":
            continue
        s = result[col].astype(str).map(clean_text)
        mask = s.eq(from_val)
        if mask.any():
            result.loc[mask, col] = to_val
    return result


def apply_advanced_conditions(df, conditions):
    result = df
    for cond in conditions:
        col = clean_text(cond.get("column"))
        mode = clean_text(cond.get("mode")).lower() or "eq"
        values = [clean_text(v) for v in cond.get("values", []) if clean_text(v)]

        if not col or col not in result.columns or not values:
            continue

        s = result[col].astype(str).map(clean_text)

        if mode == "eq":
            mask = s.isin(values)
        elif mode == "neq":
            mask = ~s.isin(values)
        elif mode == "contains":
            mask = pd.Series(False, index=result.index)
            for v in values:
                mask = mask | s.str.contains(re.escape(v), na=False, regex=True)
        elif mode == "not_contains":
            mask = pd.Series(True, index=result.index)
            for v in values:
                mask = mask & (~s.str.contains(re.escape(v), na=False, regex=True))
        elif mode == "regex":
            mask = pd.Series(False, index=result.index)
            for pattern in values:
                try:
                    mask = mask | s.str.contains(pattern, na=False, regex=True)
                except re.error:
                    continue
        elif mode in ["gt", "gte", "lt", "lte", "between"]:
            num_s = to_numeric_series(result[col])
            try:
                if mode == "gt":
                    target = float(values[0].replace(",", ""))
                    mask = num_s > target
                elif mode == "gte":
                    target = float(values[0].replace(",", ""))
                    mask = num_s >= target
                elif mode == "lt":
                    target = float(values[0].replace(",", ""))
                    mask = num_s < target
                elif mode == "lte":
                    target = float(values[0].replace(",", ""))
                    mask = num_s <= target
                else:
                    v1 = float(values[0].replace(",", ""))
                    v2 = float(values[1].replace(",", "")) if len(values) > 1 else v1
                    low, high = sorted([v1, v2])
                    mask = num_s.between(low, high, inclusive="both")
            except Exception:
                mask = pd.Series(False, index=result.index)
        elif mode in ["date_eq", "date_before", "date_after", "date_between"]:
            dt_s = to_datetime_series(result[col])
            try:
                if mode == "date_eq":
                    target = pd.to_datetime(values[0], errors="coerce")
                    mask = dt_s.dt.normalize() == target.normalize()
                elif mode == "date_before":
                    target = pd.to_datetime(values[0], errors="coerce")
                    mask = dt_s < target
                elif mode == "date_after":
                    target = pd.to_datetime(values[0], errors="coerce")
                    mask = dt_s > target
                else:
                    d1 = pd.to_datetime(values[0], errors="coerce")
                    d2 = pd.to_datetime(values[1], errors="coerce") if len(values) > 1 else d1
                    low, high = sorted([d1, d2])
                    mask = (dt_s >= low) & (dt_s <= high)
            except Exception:
                mask = pd.Series(False, index=result.index)
        else:
            mask = s.isin(values)

        result = result[mask.fillna(False)]
        if result.empty:
            break
    return result


def apply_sorting(df, sort_specs):
    if df.empty or not sort_specs:
        return df
    by = []
    ascending = []
    for spec in sort_specs:
        col = clean_text(spec.get("column"))
        order = clean_text(spec.get("order")).lower() or "asc"
        if col and col in df.columns:
            by.append(col)
            ascending.append(order != "desc")
    if not by:
        return df
    try:
        return df.sort_values(by=by, ascending=ascending, kind="stable")
    except Exception:
        return df


def apply_dedup(df, dedup_spec):
    if df.empty or not dedup_spec:
        return df
    col = clean_text(dedup_spec.get("column"))
    keep = clean_text(dedup_spec.get("keep")).lower() or "first"
    if col not in df.columns:
        return df
    try:
        return df.drop_duplicates(subset=[col], keep="last" if keep == "last" else "first")
    except Exception:
        return df


# =========================================================
# 파일 읽기
# =========================================================
def read_csv_header_fast(file_path, scan_rows=200):
    encodings = ["utf-8-sig", "cp949", "euc-kr", "utf-8"]
    last_error = None
    for enc in encodings:
        try:
            sample = pd.read_csv(file_path, nrows=scan_rows, header=None, dtype=str, encoding=enc)
            rows = sample.values.tolist()
            if not rows:
                return [], 0, enc
            header_idx = detect_header_index_from_rows(rows)
            header = normalize_columns_from_header_row(rows[header_idx])
            return header, header_idx, enc
        except Exception as e:
            last_error = e
    raise last_error


def read_xlsx_header_fast(file_path, sheet_name=None, scan_rows=200):
    wb = load_workbook(file_path, read_only=True, data_only=True)
    if sheet_name:
        if sheet_name not in wb.sheetnames:
            wb.close()
            raise ValueError(f"시트 '{sheet_name}' 이(가) 파일에 없습니다: {os.path.basename(file_path)}")
        ws = wb[sheet_name]
    else:
        ws = wb[wb.sheetnames[0]]
    rows = []
    for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
        rows.append(list(row))
        if i >= scan_rows:
            break
    wb.close()
    if not rows:
        return [], 0
    max_len = 0
    for r in rows:
        last_nonempty = 0
        for idx, v in enumerate(r, start=1):
            if clean_text(v):
                last_nonempty = idx
        max_len = max(max_len, last_nonempty)
    trimmed_rows = [r[:max_len] if max_len > 0 else r for r in rows]
    header_idx = detect_header_index_from_rows(trimmed_rows)
    header = normalize_columns_from_header_row(trimmed_rows[header_idx])
    return header, header_idx


def get_sheet_names(file_path):
    ext = os.path.splitext(file_path)[1].lower()
    if ext in [".xlsx", ".xlsm"] and not is_html_content(file_path):
        try:
            wb = load_workbook(file_path, read_only=True, data_only=True)
            names = wb.sheetnames
            wb.close()
            return names
        except Exception:
            return []
    return []


def extract_columns_fast(file_path, sheet_name=None, header_row_idx=None, force_html=False):
    ext = os.path.splitext(file_path)[1].lower()

    if ext == ".csv" and not force_html:
        header, _, _ = read_csv_header_fast(file_path, scan_rows=200)
        if header_row_idx is None:
            return header
        df = pd.read_csv(file_path, nrows=200, header=None, dtype=str, encoding=read_csv_header_fast(file_path, scan_rows=200)[2])
        rows = df.values.tolist()
        if not rows:
            return []
        header_idx = choose_header_index(rows, header_row_idx=header_row_idx, max_scan=200)
        return normalize_columns_from_header_row(rows[header_idx])

    if force_html or ext in [".html", ".htm"] or (ext == ".xls" and is_html_content(file_path)):
        tables = pd.read_html(file_path, header=None)
        if not tables:
            return []
        df = pd.concat(tables, ignore_index=True, sort=False)
        rows = df.values.tolist()
        header_idx = choose_header_index(rows[:200] if len(rows) > 200 else rows, header_row_idx=header_row_idx, max_scan=200)
        return normalize_columns_from_header_row(rows[header_idx])

    if ext in [".xlsx", ".xlsm"]:
        if header_row_idx is None:
            header, _ = read_xlsx_header_fast(file_path, sheet_name=sheet_name, scan_rows=200)
            return header
        wb = load_workbook(file_path, read_only=True, data_only=True)
        ws = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb[wb.sheetnames[0]]
        rows = []
        for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
            rows.append(list(row))
            if i >= 200:
                break
        wb.close()
        if not rows:
            return []
        max_len = 0
        for r in rows:
            last_nonempty = 0
            for idx, v in enumerate(r, start=1):
                if clean_text(v):
                    last_nonempty = idx
            max_len = max(max_len, last_nonempty)
        rows = [r[:max_len] if max_len > 0 else r for r in rows]
        header_idx = choose_header_index(rows, header_row_idx=header_row_idx, max_scan=200)
        return normalize_columns_from_header_row(rows[header_idx])

    if ext == ".xls":
        try:
            tables = pd.read_html(file_path, header=None)
            if tables:
                df = pd.concat(tables, ignore_index=True, sort=False)
                rows = df.values.tolist()
                header_idx = choose_header_index(rows[:200] if len(rows) > 200 else rows, header_row_idx=header_row_idx, max_scan=200)
                return normalize_columns_from_header_row(rows[header_idx])
        except Exception:
            pass
        raise ValueError(f"일반 구형 .xls 파일은 지원하지 않습니다: {os.path.basename(file_path)}")

    if ext == ".xlsb":
        raise ValueError(f".xlsb 파일은 지원하지 않습니다: {os.path.basename(file_path)}")

    return []


def extract_unique_values_fast(file_path, target_col, sheet_name=None, max_values=200, max_scan_rows=50000,
                               header_row_idx=None, force_html=False):
    ext = os.path.splitext(file_path)[1].lower()
    values = []

    def finalize(vals):
        cleaned = []
        seen = set()
        for v in vals:
            t = clean_text(v)
            if not t:
                continue
            if t not in seen:
                seen.add(t)
                cleaned.append(t)
        return cleaned[:max_values]

    if ext == ".csv" and not force_html:
        header, auto_header_idx, enc = read_csv_header_fast(file_path, scan_rows=200)
        header_idx = header_row_idx if header_row_idx is not None else auto_header_idx

        sample = pd.read_csv(file_path, nrows=200, header=None, dtype=str, encoding=enc)
        rows = sample.values.tolist()
        if not rows:
            return []
        header_idx = choose_header_index(rows, header_row_idx=header_row_idx, max_scan=200)
        header = normalize_columns_from_header_row(rows[header_idx])

        if target_col not in header:
            return []

        target_idx = header.index(target_col)
        reader = pd.read_csv(file_path, header=None, dtype=str, encoding=enc, chunksize=50000)
        skipped_rows = 0
        scanned_rows = 0
        for chunk in reader:
            if skipped_rows <= header_idx:
                need_skip = header_idx + 1 - skipped_rows
                if need_skip > 0:
                    if need_skip >= len(chunk):
                        skipped_rows += len(chunk)
                        continue
                    chunk = chunk.iloc[need_skip:].reset_index(drop=True)
                    skipped_rows = header_idx + 1
            if target_idx >= chunk.shape[1]:
                continue
            values.extend(chunk.iloc[:, target_idx].tolist())
            scanned_rows += len(chunk)
            if len(values) >= max_values * 3 or scanned_rows >= max_scan_rows:
                break
        return finalize(values)

    if force_html or ext in [".html", ".htm"] or (ext == ".xls" and is_html_content(file_path)):
        tables = pd.read_html(file_path, header=None)
        if not tables:
            return []
        df = pd.concat(tables, ignore_index=True, sort=False)
        rows = df.values.tolist()
        if not rows:
            return []
        header_idx = choose_header_index(rows[:200] if len(rows) > 200 else rows, header_row_idx=header_row_idx, max_scan=200)
        header = normalize_columns_from_header_row(rows[header_idx])
        if target_col not in header:
            return []
        body = df.iloc[header_idx + 1:].reset_index(drop=True).copy()
        if body.shape[1] < len(header):
            for _ in range(len(header) - body.shape[1]):
                body[body.shape[1]] = None
        elif body.shape[1] > len(header):
            body = body.iloc[:, :len(header)]
        body.columns = header
        return finalize(body[target_col].tolist())

    if ext in [".xlsx", ".xlsm"]:
        df = load_file_to_df(file_path, sheet_name=sheet_name, header_row_idx=header_row_idx)
        if df.empty or target_col not in df.columns:
            return []
        scanned = 0
        for v in df[target_col].tolist():
            values.append(v)
            scanned += 1
            if len(values) >= max_values * 3 or scanned >= max_scan_rows:
                break
        return finalize(values)

    if ext == ".xls":
        try:
            tables = pd.read_html(file_path, header=None)
            if tables:
                df = pd.concat(tables, ignore_index=True, sort=False)
                rows = df.values.tolist()
                if not rows:
                    return []
                header_idx = choose_header_index(rows[:200] if len(rows) > 200 else rows, header_row_idx=header_row_idx, max_scan=200)
                header = normalize_columns_from_header_row(rows[header_idx])
                if target_col not in header:
                    return []
                body = df.iloc[header_idx + 1:].reset_index(drop=True).copy()
                if body.shape[1] < len(header):
                    for _ in range(len(header) - body.shape[1]):
                        body[body.shape[1]] = None
                elif body.shape[1] > len(header):
                    body = body.iloc[:, :len(header)]
                body.columns = header
                return finalize(body[target_col].tolist())
        except Exception:
            pass
    return []


def load_file_to_df(file_path, sheet_name=None, header_row_idx=None, force_html=False):
    ext = os.path.splitext(file_path)[1].lower()

    if ext == ".csv" and not force_html:
        header, auto_header_idx, enc = read_csv_header_fast(file_path, scan_rows=200)
        df = pd.read_csv(file_path, header=None, dtype=str, encoding=enc)
        rows = df.values.tolist()
        if not rows:
            return pd.DataFrame()
        header_idx = choose_header_index(rows[:200] if len(rows) > 200 else rows, header_row_idx=header_row_idx, max_scan=200)
        header = normalize_columns_from_header_row(rows[header_idx])
        body = df.iloc[header_idx + 1:].reset_index(drop=True).copy()
        if body.shape[1] < len(header):
            for _ in range(len(header) - body.shape[1]):
                body[body.shape[1]] = None
        elif body.shape[1] > len(header):
            body = body.iloc[:, :len(header)]
        body.columns = header
        return trim_empty_columns_df(body)

    if force_html or ext in [".html", ".htm"] or (ext == ".xls" and is_html_content(file_path)):
        tables = pd.read_html(file_path, header=None)
        if not tables:
            return pd.DataFrame()
        df = pd.concat(tables, ignore_index=True, sort=False)
        rows = df.values.tolist()
        if not rows:
            return pd.DataFrame()
        header_idx = choose_header_index(rows[:200] if len(rows) > 200 else rows, header_row_idx=header_row_idx, max_scan=200)
        header = normalize_columns_from_header_row(rows[header_idx])
        body = df.iloc[header_idx + 1:].reset_index(drop=True).copy()
        if body.shape[1] < len(header):
            for _ in range(len(header) - body.shape[1]):
                body[body.shape[1]] = None
        elif body.shape[1] > len(header):
            body = body.iloc[:, :len(header)]
        body.columns = header
        return trim_empty_columns_df(body)

    if ext in [".xlsx", ".xlsm"]:
        wb = load_workbook(file_path, read_only=True, data_only=True)
        if sheet_name:
            if sheet_name not in wb.sheetnames:
                wb.close()
                raise ValueError(f"시트 '{sheet_name}' 이(가) 파일에 없습니다: {os.path.basename(file_path)}")
            ws = wb[sheet_name]
        else:
            ws = wb[wb.sheetnames[0]]

        rows = []
        for row in ws.iter_rows(values_only=True):
            rows.append(list(row))
        wb.close()

        return dataframe_from_rows_with_header(rows, header_row_idx=header_row_idx, scan_rows=200)

    if ext == ".xls":
        try:
            tables = pd.read_html(file_path, header=None)
            if tables:
                df = pd.concat(tables, ignore_index=True, sort=False)
                rows = df.values.tolist()
                if not rows:
                    return pd.DataFrame()
                header_idx = choose_header_index(rows[:200] if len(rows) > 200 else rows, header_row_idx=header_row_idx, max_scan=200)
                header = normalize_columns_from_header_row(rows[header_idx])
                body = df.iloc[header_idx + 1:].reset_index(drop=True).copy()
                if body.shape[1] < len(header):
                    for _ in range(len(header) - body.shape[1]):
                        body[body.shape[1]] = None
                elif body.shape[1] > len(header):
                    body = body.iloc[:, :len(header)]
                body.columns = header
                return trim_empty_columns_df(body)
        except Exception:
            pass

    raise ValueError(f"지원하지 않는 형식: {os.path.basename(file_path)}")


def export_from_df(df, output_path, selected_columns=None, conditions=None, replacement_rules=None,
                   fill_service=True, sort_specs=None, dedup_spec=None):
    if replacement_rules:
        df = apply_replacements(df, replacement_rules)
    if conditions:
        df = apply_advanced_conditions(df, conditions)
    if fill_service:
        df = fill_service_small_from_mid(df)
    if selected_columns:
        keep_cols = [c for c in selected_columns if c in df.columns]
        df = df[keep_cols]
    if sort_specs:
        df = apply_sorting(df, sort_specs)
    if dedup_spec:
        df = apply_dedup(df, dedup_spec)
    df.to_csv(output_path, index=False, encoding="utf-8-sig")
    return len(df)


# =========================================================
# 열려있는 엑셀
# =========================================================
def list_open_excel_workbooks():
    if not WIN32_AVAILABLE:
        return []
    try:
        app = win32com.client.GetActiveObject("Excel.Application")
        items = []
        for wb in app.Workbooks:
            try:
                full_name = wb.FullName
            except Exception:
                full_name = ""
            items.append({
                "name": clean_text(wb.Name),
                "full_name": clean_text(full_name),
                "saved": bool(getattr(wb, "Saved", False)),
                "sheet_names": [clean_text(ws.Name) for ws in wb.Worksheets]
            })
        return items
    except Exception:
        return []


def export_open_workbook_sheet_to_temp(workbook_name, sheet_name):
    if not WIN32_AVAILABLE:
        raise RuntimeError("pywin32가 설치되지 않았습니다.")

    pythoncom.CoInitialize()
    app = win32com.client.GetActiveObject("Excel.Application")
    target_wb = None

    for wb in app.Workbooks:
        if clean_text(wb.Name) == clean_text(workbook_name):
            target_wb = wb
            break

    if target_wb is None:
        pythoncom.CoUninitialize()
        raise RuntimeError("열려있는 워크북을 찾지 못했습니다.")

    if not clean_text(sheet_name):
        pythoncom.CoUninitialize()
        raise RuntimeError("시트가 선택되지 않았습니다.")

    tmp_dir = os.path.join(tempfile.gettempdir(), "gui_data_extractor_open_excel")
    os.makedirs(tmp_dir, exist_ok=True)
    tmp_path = os.path.join(tmp_dir, f"{uuid.uuid4().hex}.xlsx")

    new_wb = None
    prev_alerts = getattr(app, "DisplayAlerts", True)
    prev_screen = getattr(app, "ScreenUpdating", True)
    try:
        app.DisplayAlerts = False
        app.ScreenUpdating = False
        sheet = target_wb.Worksheets(sheet_name)
        sheet.Copy()
        new_wb = app.ActiveWorkbook
        xlOpenXMLWorkbook = 51
        new_wb.SaveAs(tmp_path, FileFormat=xlOpenXMLWorkbook)
        new_wb.Close(SaveChanges=False)
        return tmp_path, None
    except Exception as e:
        try:
            if new_wb is not None:
                new_wb.Close(SaveChanges=False)
        except Exception:
            pass
        raise RuntimeError(f"열린 시트 임시 저장 실패: {e}")
    finally:
        try:
            app.DisplayAlerts = prev_alerts
            app.ScreenUpdating = prev_screen
        except Exception:
            pass
        pythoncom.CoUninitialize()


def _normalize_excel_value_matrix(values):
    if values is None:
        return []

    if not isinstance(values, tuple):
        return [[values]]

    rows = []
    for row in values:
        if isinstance(row, tuple):
            rows.append(list(row))
        else:
            rows.append([row])
    return rows


def _trim_rows_to_used_content(rows):
    if not rows:
        return rows

    max_len = 0
    for r in rows:
        last_nonempty = 0
        for idx, v in enumerate(r, start=1):
            if clean_text(v):
                last_nonempty = idx
        max_len = max(max_len, last_nonempty)

    if max_len <= 0:
        return rows

    return [list(r[:max_len]) for r in rows]


def read_open_excel_sheet_rows(workbook_name, sheet_name, max_rows=None):
    if not WIN32_AVAILABLE:
        raise RuntimeError("pywin32가 설치되지 않았습니다.")

    pythoncom.CoInitialize()
    try:
        app = win32com.client.GetActiveObject("Excel.Application")
        target_wb = None
        for wb in app.Workbooks:
            if clean_text(wb.Name) == clean_text(workbook_name):
                target_wb = wb
                break

        if target_wb is None:
            raise RuntimeError("열려있는 워크북을 찾지 못했습니다.")

        ws = target_wb.Worksheets(sheet_name)
        used = ws.UsedRange
        values = used.Value
        rows = _normalize_excel_value_matrix(values)
        rows = _trim_rows_to_used_content(rows)

        if max_rows is not None and max_rows > 0:
            rows = rows[:max_rows]

        return rows
    finally:
        pythoncom.CoUninitialize()


def extract_open_excel_columns(workbook_name, sheet_name, scan_rows=200, header_row_idx=None):
    rows = read_open_excel_sheet_rows(workbook_name, sheet_name, max_rows=max(scan_rows, 1))
    if not rows:
        return []
    header_idx = choose_header_index(rows[:scan_rows] if len(rows) > scan_rows else rows, header_row_idx=header_row_idx, max_scan=scan_rows)
    return normalize_columns_from_header_row(rows[header_idx])


def load_open_excel_sheet_df(workbook_name, sheet_name, max_rows=None, header_row_idx=None):
    rows = read_open_excel_sheet_rows(workbook_name, sheet_name, max_rows=max_rows)
    if not rows:
        return pd.DataFrame()
    return dataframe_from_rows_with_header(rows, header_row_idx=header_row_idx, scan_rows=200)


def extract_open_excel_unique_values(workbook_name, sheet_name, target_col, max_values=200, max_scan_rows=50000, header_row_idx=None):
    df = load_open_excel_sheet_df(workbook_name, sheet_name, max_rows=max_scan_rows, header_row_idx=header_row_idx)
    if df.empty or target_col not in df.columns:
        return []
    cleaned = []
    seen = set()
    for v in df[target_col].tolist():
        t = clean_text(v)
        if not t or t in seen:
            continue
        seen.add(t)
        cleaned.append(t)
        if len(cleaned) >= max_values:
            break
    return cleaned


# =========================================================
# 워커
# =========================================================
class MergeWorker(QThread):
    progress_changed = Signal(int)
    status_changed = Signal(str)
    finished_ok = Signal(str)
    error_occurred = Signal(str)

    def __init__(self, files, output_path, sheet_name=None, selected_columns=None,
                 conditions=None, replacement_rules=None, fill_service=True):
        super().__init__()
        self.files = files
        self.output_path = output_path
        self.sheet_name = sheet_name
        self.selected_columns = selected_columns or []
        self.conditions = conditions or []
        self.replacement_rules = replacement_rules or []
        self.fill_service = fill_service

    def run(self):
        try:
            result_frames = []
            total_files = len(self.files)
            failed = []

            for idx, file_path in enumerate(self.files, start=1):
                base = os.path.basename(file_path)
                try:
                    if is_file_locked(file_path):
                        self.status_changed.emit(f"[열림 감지] {base} / 임시 복사본으로 처리")
                    readable = get_readable_file_path(file_path)
                    df = load_file_to_df(readable, sheet_name=self.sheet_name)
                    if self.replacement_rules:
                        df = apply_replacements(df, self.replacement_rules)
                    if self.conditions:
                        df = apply_advanced_conditions(df, self.conditions)
                    if self.fill_service:
                        df = fill_service_small_from_mid(df)
                    if self.selected_columns:
                        keep = [c for c in self.selected_columns if c in df.columns]
                        df = df[keep]
                    result_frames.append(df)
                    self.status_changed.emit(f"[완료] {base} / {len(df):,}행")
                except Exception as e:
                    failed.append((base, str(e)))
                    self.status_changed.emit(f"[실패] {base} / {e}")
                self.progress_changed.emit(int(idx / total_files * 100) if total_files else 100)

            final_df = pd.concat(result_frames, ignore_index=True) if result_frames else pd.DataFrame()
            final_df.to_csv(self.output_path, index=False, encoding="utf-8-sig")

            msg = [f"병합 완료: 총 {len(final_df):,}행 저장", f"성공 파일: {len(result_frames)}개", f"실패 파일: {len(failed)}개"]
            if failed:
                msg.append("")
                msg.append("[실패 파일 목록]")
                for name, reason in failed[:20]:
                    msg.append(f"- {name}: {reason}")
            self.finished_ok.emit("\n".join(msg))
        except Exception as e:
            self.error_occurred.emit(f"{e}\n\n{traceback.format_exc()}")


class ExportWorker(QThread):
    progress_changed = Signal(int)
    status_changed = Signal(str)
    finished_ok = Signal(str)
    error_occurred = Signal(str)

    def __init__(self, source_path, output_path, sheet_name=None, selected_columns=None,
                 conditions=None, replacement_rules=None, fill_service=True,
                 sort_specs=None, dedup_spec=None, header_row_idx=None, force_html=False):
        super().__init__()
        self.source_path = source_path
        self.output_path = output_path
        self.sheet_name = sheet_name
        self.selected_columns = selected_columns or []
        self.conditions = conditions or []
        self.replacement_rules = replacement_rules or []
        self.fill_service = fill_service
        self.sort_specs = sort_specs or []
        self.dedup_spec = dedup_spec
        self.header_row_idx = header_row_idx
        self.force_html = force_html

    def run(self):
        try:
            self.progress_changed.emit(10)
            readable = get_readable_file_path(self.source_path)
            self.status_changed.emit(f"[읽기] {os.path.basename(self.source_path)}")
            df = load_file_to_df(
                readable,
                sheet_name=self.sheet_name,
                header_row_idx=self.header_row_idx,
                force_html=self.force_html
            )
            self.progress_changed.emit(50)
            written = export_from_df(
                df=df,
                output_path=self.output_path,
                selected_columns=self.selected_columns,
                conditions=self.conditions,
                replacement_rules=self.replacement_rules,
                fill_service=self.fill_service,
                sort_specs=self.sort_specs,
                dedup_spec=self.dedup_spec,
            )
            self.progress_changed.emit(100)
            self.finished_ok.emit(f"완료: 총 {written:,}행 저장")
        except Exception as e:
            self.error_occurred.emit(f"{e}\n\n{traceback.format_exc()}")


class OpenExcelExportWorker(QThread):
    progress_changed = Signal(int)
    status_changed = Signal(str)
    finished_ok = Signal(str)
    error_occurred = Signal(str)

    def __init__(self, workbook_name, sheet_name, output_path, selected_columns=None,
                 conditions=None, replacement_rules=None, fill_service=True,
                 sort_specs=None, dedup_spec=None, header_row_idx=None):
        super().__init__()
        self.workbook_name = workbook_name
        self.sheet_name = sheet_name
        self.output_path = output_path
        self.selected_columns = selected_columns or []
        self.conditions = conditions or []
        self.replacement_rules = replacement_rules or []
        self.fill_service = fill_service
        self.sort_specs = sort_specs or []
        self.dedup_spec = dedup_spec
        self.header_row_idx = header_row_idx

    def run(self):
        try:
            self.progress_changed.emit(10)
            self.status_changed.emit(f"[읽기] {self.workbook_name} / {self.sheet_name}")
            df = load_open_excel_sheet_df(
                self.workbook_name,
                self.sheet_name,
                max_rows=None,
                header_row_idx=self.header_row_idx
            )
            self.progress_changed.emit(55)
            written = export_from_df(
                df=df,
                output_path=self.output_path,
                selected_columns=self.selected_columns,
                conditions=self.conditions,
                replacement_rules=self.replacement_rules,
                fill_service=self.fill_service,
                sort_specs=self.sort_specs,
                dedup_spec=self.dedup_spec,
            )
            self.progress_changed.emit(100)
            self.finished_ok.emit(f"완료: 총 {written:,}행 저장")
        except Exception as e:
            self.error_occurred.emit(f"{e}\n\n{traceback.format_exc()}")

# =========================================================
# UI 위젯
# =========================================================
class ColumnCheckGrid(QWidget):
    def __init__(self, title="출력 컬럼 선택", max_columns=5):
        super().__init__()
        self.all_columns = []
        self.checkboxes = {}
        self.max_columns = max_columns

        root = QVBoxLayout(self)
        top = QHBoxLayout()
        self.lbl_title = QLabel(title)
        top.addWidget(self.lbl_title)
        top.addStretch()
        self.lbl_count = QLabel("0 / 0 선택")
        top.addWidget(self.lbl_count)
        root.addLayout(top)

        search_row = QHBoxLayout()
        self.search = QLineEdit()
        self.search.setPlaceholderText("컬럼 검색")
        self.search.textChanged.connect(self.rebuild)
        search_row.addWidget(self.search)
        self.btn_all = QPushButton("전체 선택")
        self.btn_all.clicked.connect(self.check_all)
        search_row.addWidget(self.btn_all)
        self.btn_none = QPushButton("전체 해제")
        self.btn_none.clicked.connect(self.uncheck_all)
        search_row.addWidget(self.btn_none)
        root.addLayout(search_row)

        self.scroll = QScrollArea()
        self.scroll.setWidgetResizable(True)
        self.inner = QWidget()
        self.grid = QGridLayout(self.inner)
        self.grid.setContentsMargins(8, 8, 8, 8)
        self.grid.setHorizontalSpacing(12)
        self.grid.setVerticalSpacing(8)
        self.scroll.setWidget(self.inner)
        root.addWidget(self.scroll)

    def resizeEvent(self, event):
        super().resizeEvent(event)
        self.rebuild()

    def set_columns(self, columns, checked=True, preserve_checked=None):
        prev_checked = set(preserve_checked or [])
        self.all_columns = list(columns)
        self.checkboxes = {}
        for col in self.all_columns:
            cb = QCheckBox(str(col))
            cb.setChecked((col in prev_checked) if prev_checked else checked)
            cb.stateChanged.connect(self.update_count)
            self.checkboxes[col] = cb
        self.rebuild()

    def visible_column_count(self):
        return 4

    def rebuild(self):
        while self.grid.count():
            item = self.grid.takeAt(0)
            w = item.widget()
            if w:
                w.setParent(None)
        keyword = clean_text(self.search.text()).lower()
        columns = self.all_columns
        if keyword:
            columns = [c for c in columns if keyword in str(c).lower()]
        col_count = self.visible_column_count()
        for idx, col in enumerate(columns):
            row = idx // col_count
            col_idx = idx % col_count
            cb = self.checkboxes[col]
            cb.setMinimumWidth(0)
            cb.setMaximumWidth(220)
            self.grid.addWidget(cb, row, col_idx)
        for i in range(col_count):
            self.grid.setColumnStretch(i, 0)
        self.grid.setColumnStretch(col_count, 1)
        self.update_count()

    def update_count(self):
        total = len(self.all_columns)
        selected = len([c for c, cb in self.checkboxes.items() if cb.isChecked()])
        self.lbl_count.setText(f"{selected} / {total} 선택")

    def get_checked_columns(self):
        return [c for c, cb in self.checkboxes.items() if cb.isChecked()]

    def check_all(self):
        for cb in self.checkboxes.values():
            cb.setChecked(True)
        self.update_count()

    def uncheck_all(self):
        for cb in self.checkboxes.values():
            cb.setChecked(False)
        self.update_count()


class ValueFilterPanel(QWidget):
    def __init__(self, title="실제 값 선택", max_columns=3):
        super().__init__()
        self.all_values = []
        self.checkboxes = {}
        self.max_columns = max_columns
        root = QVBoxLayout(self)
        top = QHBoxLayout()
        self.lbl_title = QLabel(title)
        top.addWidget(self.lbl_title)
        top.addStretch()
        self.lbl_count = QLabel("0 / 0 선택")
        top.addWidget(self.lbl_count)
        root.addLayout(top)
        search_row = QHBoxLayout()
        self.search = QLineEdit()
        self.search.setPlaceholderText("값 검색")
        self.search.textChanged.connect(self.rebuild)
        search_row.addWidget(self.search)
        self.btn_all = QPushButton("전체 선택")
        self.btn_all.clicked.connect(self.check_all)
        search_row.addWidget(self.btn_all)
        self.btn_none = QPushButton("전체 해제")
        self.btn_none.clicked.connect(self.uncheck_all)
        search_row.addWidget(self.btn_none)
        root.addLayout(search_row)
        self.scroll = QScrollArea()
        self.scroll.setWidgetResizable(True)
        self.inner = QWidget()
        self.grid = QGridLayout(self.inner)
        self.grid.setContentsMargins(8, 8, 8, 8)
        self.grid.setHorizontalSpacing(12)
        self.grid.setVerticalSpacing(8)
        self.scroll.setWidget(self.inner)
        root.addWidget(self.scroll)

    def resizeEvent(self, event):
        super().resizeEvent(event)
        self.rebuild()

    def set_values(self, values, checked=False):
        self.all_values = list(values)[:200]
        self.checkboxes = {}
        for v in self.all_values:
            cb = QCheckBox(str(v))
            cb.setChecked(checked)
            cb.stateChanged.connect(self.update_count)
            self.checkboxes[v] = cb
        self.rebuild()

    def visible_column_count(self):
        return 3

    def rebuild(self):
        while self.grid.count():
            item = self.grid.takeAt(0)
            w = item.widget()
            if w:
                w.setParent(None)
        keyword = clean_text(self.search.text()).lower()
        values = self.all_values
        if keyword:
            values = [v for v in values if keyword in str(v).lower()]
        col_count = self.visible_column_count()
        for idx, v in enumerate(values):
            row = idx // col_count
            col_idx = idx % col_count
            cb = self.checkboxes[v]
            cb.setMinimumWidth(0)
            cb.setMaximumWidth(260)
            self.grid.addWidget(cb, row, col_idx)
        for i in range(col_count):
            self.grid.setColumnStretch(i, 0)
        self.grid.setColumnStretch(col_count, 1)
        self.update_count()

    def update_count(self):
        total = len(self.all_values)
        selected = len([v for v, cb in self.checkboxes.items() if cb.isChecked()])
        self.lbl_count.setText(f"{selected} / {total} 선택")

    def get_checked_values(self):
        return [v for v, cb in self.checkboxes.items() if cb.isChecked()]

    def check_all(self):
        for cb in self.checkboxes.values():
            cb.setChecked(True)
        self.update_count()

    def uncheck_all(self):
        for cb in self.checkboxes.values():
            cb.setChecked(False)
        self.update_count()



class ConditionEditor(QWidget):
    MODE_ITEMS = [
        ("같음 (=)", "eq", "값이 정확히 같은 행만 추출합니다.", "예: 강원본부"),
        ("같지 않음 (제외)", "neq", "입력한 값을 제외하고 추출합니다.", "예: 없음"),
        ("포함 (부분 일치)", "contains", "입력한 글자가 포함된 행을 추출합니다.", "예: 김, 장애"),
        ("포함 안함", "not_contains", "입력한 글자가 포함된 행을 제외합니다.", "예: 해지, 중지"),
        ("패턴 검색 (정규식)", "regex", "고급 패턴으로 찾습니다. 익숙할 때만 사용하세요.", r"예: ^G\\d{3}"),
        ("숫자 초과 (>)", "gt", "입력값보다 큰 숫자만 추출합니다.", "예: 100000"),
        ("숫자 이상 (>=)", "gte", "입력값 이상인 숫자만 추출합니다.", "예: 100000"),
        ("숫자 미만 (<)", "lt", "입력값보다 작은 숫자만 추출합니다.", "예: 100000"),
        ("숫자 이하 (<=)", "lte", "입력값 이하인 숫자만 추출합니다.", "예: 100000"),
        ("숫자 범위 (A~B)", "between", "두 숫자 사이의 범위를 찾습니다.", "예: 1000,5000"),
        ("날짜와 같음", "date_eq", "입력한 날짜와 같은 행만 추출합니다.", "예: 2026-04-21"),
        ("날짜 이전", "date_before", "입력한 날짜보다 이전인 행을 추출합니다.", "예: 2026-04-21"),
        ("날짜 이후", "date_after", "입력한 날짜보다 이후인 행을 추출합니다.", "예: 2026-04-21"),
        ("날짜 범위", "date_between", "두 날짜 사이의 범위를 찾습니다.", "예: 2026-01-01,2026-03-31"),
    ]
    DISPLAY_TO_CODE = {display: code for display, code, _, _ in MODE_ITEMS}
    CODE_TO_DISPLAY = {code: display for display, code, _, _ in MODE_ITEMS}
    CODE_TO_DESC = {code: desc for _, code, desc, _ in MODE_ITEMS}
    CODE_TO_EXAMPLE = {code: example for _, code, _, example in MODE_ITEMS}

    def __init__(self, title="조건 설정"):
        super().__init__()
        self.source_path_getter = None
        self.sheet_name_getter = None
        self.custom_values_getter = None
        self.value_cache = {}

        root = QVBoxLayout(self)

        top_title = QLabel(title)
        top_title.setObjectName("sectionTitle")
        root.addWidget(top_title)

        help_box = QFrame()
        help_box.setObjectName("helpCard")
        help_l = QVBoxLayout(help_box)
        help_l.setContentsMargins(10, 10, 10, 10)

        help_title = QLabel("조건 선택 도움말")
        help_title.setObjectName("helpTitle")
        help_l.addWidget(help_title)

        self.lbl_mode_desc = QLabel("조건 종류를 고르면 뜻과 예시가 여기 표시됩니다.")
        self.lbl_mode_desc.setWordWrap(True)
        help_l.addWidget(self.lbl_mode_desc)
        root.addWidget(help_box)

        row = QHBoxLayout()
        self.cmb_col = QComboBox()
        self.cmb_col.currentIndexChanged.connect(self.refresh_values)
        row.addWidget(self.cmb_col, 3)

        self.cmb_mode = QComboBox()
        self.cmb_mode.addItems([display for display, _, _, _ in self.MODE_ITEMS])
        self.cmb_mode.currentIndexChanged.connect(self.update_mode_help)
        row.addWidget(self.cmb_mode, 2)

        self.edt_values = QLineEdit()
        self.edt_values.setPlaceholderText("값 입력")
        row.addWidget(self.edt_values, 4)

        self.btn_add = QPushButton("조건 추가")
        self.btn_add.clicked.connect(self.add_row_from_ui)
        row.addWidget(self.btn_add, 1)
        root.addLayout(row)

        self.value_panel = ValueFilterPanel("선택 컬럼의 실제 값", max_columns=2)
        root.addWidget(self.value_panel, 2)

        self.table = QTableWidget(0, 5)
        self.table.setHorizontalHeaderLabels(["사용", "컬럼", "조건 종류", "값", "설명"])
        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.table.verticalHeader().setVisible(False)
        self.table.setSelectionBehavior(QTableWidget.SelectRows)
        self.table.setEditTriggers(QTableWidget.DoubleClicked | QTableWidget.SelectedClicked)
        root.addWidget(self.table, 3)

        btns = QHBoxLayout()
        self.btn_remove = QPushButton("선택 조건 삭제")
        self.btn_remove.clicked.connect(self.remove_selected_rows)
        btns.addWidget(self.btn_remove)
        root.addLayout(btns)

        self.update_mode_help()

    def set_source_getters(self, path_getter, sheet_getter=None):
        self.source_path_getter = path_getter
        self.sheet_name_getter = sheet_getter

    def set_custom_values_getter(self, getter):
        self.custom_values_getter = getter

    def set_columns(self, columns):
        self.cmb_col.blockSignals(True)
        self.cmb_col.clear()
        self.cmb_col.addItems(columns)
        self.cmb_col.blockSignals(False)
        if columns:
            self.refresh_values()

    def get_mode_code(self):
        return self.DISPLAY_TO_CODE.get(clean_text(self.cmb_mode.currentText()), "eq")

    def update_mode_help(self):
        code = self.get_mode_code()
        desc = self.CODE_TO_DESC.get(code, "")
        example = self.CODE_TO_EXAMPLE.get(code, "")
        self.lbl_mode_desc.setText(f"{desc}\n{example}")
        if code in ("between", "date_between"):
            self.edt_values.setPlaceholderText("두 값을 쉼표로 입력하세요. 예: 1000,5000")
        elif code in ("gt", "gte", "lt", "lte"):
            self.edt_values.setPlaceholderText("숫자를 입력하세요. 예: 100000")
        elif code.startswith("date_"):
            self.edt_values.setPlaceholderText("날짜를 입력하세요. 예: 2026-04-21")
        elif code == "regex":
            self.edt_values.setPlaceholderText(r"패턴 입력. 예: ^G\d{3}")
        else:
            self.edt_values.setPlaceholderText("값 직접 입력 (쉼표 구분)")

    def refresh_values(self):
        col = clean_text(self.cmb_col.currentText())
        if not col:
            self.value_panel.set_values([], checked=False)
            return

        try:
            if self.custom_values_getter is not None:
                cache_key = ("custom", col)
                if cache_key in self.value_cache:
                    values = self.value_cache[cache_key]
                else:
                    values = self.custom_values_getter(col) or []
                    self.value_cache[cache_key] = values
                self.value_panel.set_values(values, checked=False)
                return

            if not self.source_path_getter:
                self.value_panel.set_values([], checked=False)
                return
            source_path = self.source_path_getter()
            if not source_path:
                self.value_panel.set_values([], checked=False)
                return
            sheet_name = self.sheet_name_getter() if self.sheet_name_getter else None
            cache_key = (source_path, sheet_name, col)
            if cache_key in self.value_cache:
                values = self.value_cache[cache_key]
            else:
                values = extract_unique_values_fast(
                    source_path,
                    col,
                    sheet_name=sheet_name,
                    max_values=200,
                    max_scan_rows=50000
                )
                self.value_cache[cache_key] = values
            self.value_panel.set_values(values, checked=False)
        except Exception:
            self.value_panel.set_values([], checked=False)

    def append_row(self, use=True, column="", mode="eq", values="", desc=""):
        row = self.table.rowCount()
        self.table.insertRow(row)
        chk = QTableWidgetItem()
        chk.setFlags(chk.flags() | Qt.ItemIsUserCheckable | Qt.ItemIsEnabled)
        chk.setCheckState(Qt.Checked if use else Qt.Unchecked)
        self.table.setItem(row, 0, chk)
        self.table.setItem(row, 1, QTableWidgetItem(str(column)))
        self.table.setItem(row, 2, QTableWidgetItem(str(self.CODE_TO_DISPLAY.get(mode, mode))))
        self.table.setItem(row, 3, QTableWidgetItem(str(values)))
        self.table.setItem(row, 4, QTableWidgetItem(str(desc or self.CODE_TO_DESC.get(mode, ""))))

    def add_row_from_ui(self):
        col = clean_text(self.cmb_col.currentText())
        mode = self.get_mode_code()
        manual_text = clean_text(self.edt_values.text())
        values = [clean_text(x) for x in manual_text.split(",") if clean_text(x)] if manual_text else self.value_panel.get_checked_values()
        if not col:
            QMessageBox.warning(self, "알림", "조건 컬럼을 선택하세요.")
            return
        if not values:
            QMessageBox.warning(self, "알림", "조건값을 입력하거나 선택하세요.")
            return
        self.append_row(True, col, mode, ",".join(values), self.CODE_TO_DESC.get(mode, ""))
        self.edt_values.clear()

    def remove_selected_rows(self):
        selected = self.table.selectionModel().selectedRows()
        if not selected:
            QMessageBox.warning(self, "알림", "삭제할 조건 행을 선택하세요.")
            return
        rows = sorted([x.row() for x in selected], reverse=True)
        for row in rows:
            self.table.removeRow(row)

    def parse_conditions(self):
        result = []
        for row in range(self.table.rowCount()):
            use_item = self.table.item(row, 0)
            if not use_item or use_item.checkState() != Qt.Checked:
                continue
            col_item = self.table.item(row, 1)
            mode_item = self.table.item(row, 2)
            val_item = self.table.item(row, 3)
            col = clean_text(col_item.text()) if col_item else ""
            mode_display = clean_text(mode_item.text()) if mode_item else "같음 (=)"
            mode = self.DISPLAY_TO_CODE.get(mode_display, mode_display.lower() or "eq")
            raw_values = clean_text(val_item.text()) if val_item else ""
            values = [clean_text(x) for x in raw_values.split(",") if clean_text(x)]
            if not col or not values:
                continue
            result.append({"column": col, "mode": mode or "eq", "values": values})
        return result

    def load_conditions(self, conditions):
        self.table.setRowCount(0)
        for cond in conditions:
            col = clean_text(cond.get("column"))
            mode = clean_text(cond.get("mode")).lower() or "eq"
            values = [clean_text(v) for v in cond.get("values", []) if clean_text(v)]
            self.append_row(True, col, mode, ",".join(values), self.CODE_TO_DESC.get(mode, ""))


class MergeTab(QWidget):

    def __init__(self, log_cb):
        super().__init__()
        self.log_cb = log_cb
        self.files = []
        self.current_folder = ""
        self.current_common_columns = []
        self.build_ui()

    def build_ui(self):
        root = QVBoxLayout(self)
        top = QHBoxLayout()
        self.btn_select_folder = QPushButton("병합 폴더 선택")
        self.btn_select_folder.clicked.connect(self.select_folder)
        top.addWidget(self.btn_select_folder)
        self.lbl_folder = QLabel("폴더를 선택하세요")
        top.addWidget(self.lbl_folder)
        top.addWidget(QLabel("시트"))
        self.cmb_sheet = QComboBox()
        self.cmb_sheet.currentIndexChanged.connect(self.refresh_common_columns)
        top.addWidget(self.cmb_sheet)
        self.chk_fill_service = QCheckBox("서비스(소) 자동 채움")
        self.chk_fill_service.setChecked(True)
        top.addWidget(self.chk_fill_service)
        top.addStretch()
        self.btn_merge = QPushButton("병합 실행")
        self.btn_merge.clicked.connect(self.run_merge)
        top.addWidget(self.btn_merge)
        root.addLayout(top)

        split = QSplitter(Qt.Horizontal)
        left = make_card(); left_l = QVBoxLayout(left)
        left_l.addWidget(QLabel("DB 파일 목록"))
        self.list_files = QListWidget()
        self.list_files.itemChanged.connect(self.refresh_common_columns)
        left_l.addWidget(self.list_files)
        self.btn_refresh_common = QPushButton("공통 컬럼 재계산")
        self.btn_refresh_common.clicked.connect(self.refresh_common_columns)
        left_l.addWidget(self.btn_refresh_common)
        split.addWidget(left)

        center = make_card(); center_l = QVBoxLayout(center)
        self.col_grid = ColumnCheckGrid("선택된 DB의 공통 컬럼", max_columns=4)
        center_l.addWidget(self.col_grid)
        split.addWidget(center)

        right = make_card(); right_l = QVBoxLayout(right)
        self.condition_editor = ConditionEditor("병합 조건")
        self.condition_editor.set_source_getters(self.get_first_readable_selected_file, self.get_selected_sheet_name)
        right_l.addWidget(self.condition_editor)
        split.addWidget(right)

        split.setStretchFactor(0, 2)
        split.setStretchFactor(1, 3)
        split.setStretchFactor(2, 4)
        root.addWidget(split, 4)

        preview_card = make_card(); preview_l = QVBoxLayout(preview_card)
        preview_l.addWidget(QLabel("미리보기"))
        self.txt_preview = QTextEdit(); self.txt_preview.setReadOnly(True)
        preview_l.addWidget(self.txt_preview)
        root.addWidget(preview_card, 2)

    def log(self, msg):
        self.log_cb(f"[병합] {msg}")

    def select_folder(self):
        folder = QFileDialog.getExistingDirectory(self, "병합 폴더 선택")
        if not folder:
            return
        self.current_folder = folder
        patterns = ["*.csv", "*.xlsx", "*.xlsm", "*.xls", "*.html", "*.htm"]
        files = []
        for p in patterns:
            files.extend(glob.glob(os.path.join(folder, "**", p), recursive=True))
        self.files = sorted(set(files))
        self.lbl_folder.setText(f"{folder} / {len(self.files)}개")
        self.list_files.clear()
        for f in self.files:
            label = os.path.relpath(f, folder)
            if is_file_locked(f):
                label += "  [열려 있음]"
            item = QListWidgetItem(label)
            item.setFlags(item.flags() | Qt.ItemIsUserCheckable)
            item.setCheckState(Qt.Checked)
            self.list_files.addItem(item)
        self.cmb_sheet.clear()
        if self.files:
            sample = get_readable_file_path(self.files[0])
            names = get_sheet_names(sample)
            if names:
                self.cmb_sheet.addItems(names)
            else:
                self.cmb_sheet.addItem("(기본)")
        self.refresh_common_columns()

    def get_selected_files(self):
        selected = []
        for i in range(self.list_files.count()):
            item = self.list_files.item(i)
            if item.checkState() == Qt.Checked:
                selected.append(self.files[i])
        return selected

    def get_selected_sheet_name(self):
        t = self.cmb_sheet.currentText()
        return None if t == "(기본)" else clean_text(t)

    def get_first_readable_selected_file(self):
        files = self.get_selected_files() or self.files
        if not files:
            return None
        return get_readable_file_path(files[0])

    def refresh_common_columns(self):
        files = self.get_selected_files() or self.files
        if not files:
            self.col_grid.set_columns([], checked=False)
            self.condition_editor.set_columns([])
            self.txt_preview.setPlainText("")
            return
        sheet_name = self.get_selected_sheet_name()
        common = None
        failed = 0
        for f in files:
            try:
                cols = set(extract_columns_fast(get_readable_file_path(f), sheet_name=sheet_name))
                common = cols if common is None else common & cols
            except Exception:
                failed += 1
        common_cols = sorted(common) if common else []
        self.current_common_columns = common_cols
        prev = set(self.col_grid.get_checked_columns())
        self.col_grid.set_columns(common_cols, checked=True, preserve_checked=prev)
        self.condition_editor.set_columns(common_cols)
        if files:
            try:
                df = load_file_to_df(get_readable_file_path(files[0]), sheet_name=sheet_name)
                if common_cols:
                    self.txt_preview.setPlainText(preview_text(df[[c for c in common_cols if c in df.columns]]))
                else:
                    self.txt_preview.setPlainText("(공통 컬럼 없음)")
            except Exception as e:
                self.txt_preview.setPlainText(f"미리보기 오류: {e}")
        if failed:
            self.log(f"공통 컬럼 계산 중 실패 파일 {failed}개")

    def run_merge(self):
        files = self.get_selected_files()
        if not files:
            QMessageBox.warning(self, "알림", "병합할 파일을 선택하세요.")
            return
        selected_columns = self.col_grid.get_checked_columns()
        if not selected_columns:
            QMessageBox.warning(self, "알림", "병합할 공통 컬럼을 선택하세요.")
            return
        output_path, _ = QFileDialog.getSaveFileName(self, "저장 파일", "병합결과.csv", "CSV Files (*.csv)")
        if not output_path:
            return
        self.worker = MergeWorker(
            files=files,
            output_path=output_path,
            sheet_name=self.get_selected_sheet_name(),
            selected_columns=selected_columns,
            conditions=self.condition_editor.parse_conditions(),
            fill_service=self.chk_fill_service.isChecked(),
        )
        self.worker.progress_changed.connect(self.parent().parent().on_progress if hasattr(self.parent().parent(), 'on_progress') else lambda x: None)
        self.worker.status_changed.connect(self.log_cb)
        self.worker.finished_ok.connect(lambda msg: QMessageBox.information(self, "완료", msg))
        self.worker.error_occurred.connect(lambda msg: QMessageBox.critical(self, "오류", msg))
        self.worker.start()


class SingleFileTab(QWidget):
    def __init__(self, log_cb):
        super().__init__()
        self.log_cb = log_cb
        self.file_path = ""
        self.build_ui()

    def build_ui(self):
        root = QVBoxLayout(self)
        top = QHBoxLayout()
        self.btn_open = QPushButton("단일 파일 열기")
        self.btn_open.clicked.connect(self.select_file)
        top.addWidget(self.btn_open)

        self.lbl_file = QLabel("파일을 선택하세요")
        top.addWidget(self.lbl_file)

        top.addWidget(QLabel("시트"))
        self.cmb_sheet = QComboBox()
        self.cmb_sheet.currentIndexChanged.connect(self.refresh_all)
        top.addWidget(self.cmb_sheet)

        top.addWidget(QLabel("헤더"))
        self.cmb_header = QComboBox()
        self.cmb_header.addItem("자동", None)
        for i in range(1, 11):
            self.cmb_header.addItem(f"{i}행", i - 1)
        self.cmb_header.currentIndexChanged.connect(self.refresh_all)
        top.addWidget(self.cmb_header)

        self.chk_force_html = QCheckBox("HTML 강제 읽기")
        self.chk_force_html.stateChanged.connect(self.refresh_all)
        top.addWidget(self.chk_force_html)

        self.chk_fill_service = QCheckBox("서비스(소) 자동 채움")
        self.chk_fill_service.setChecked(True)
        top.addWidget(self.chk_fill_service)

        self.btn_diag = QPushButton("원본 구조 보기")
        self.btn_diag.clicked.connect(self.show_raw_structure)
        top.addWidget(self.btn_diag)

        top.addStretch()

        self.btn_export = QPushButton("추출 실행")
        self.btn_export.clicked.connect(self.run_export)
        top.addWidget(self.btn_export)
        root.addLayout(top)

        split = QSplitter(Qt.Horizontal)

        left = make_card(); left_l = QVBoxLayout(left)
        self.col_grid = ColumnCheckGrid("출력 컬럼 선택", max_columns=4)
        left_l.addWidget(self.col_grid)
        split.addWidget(left)

        center = make_card(); center_l = QVBoxLayout(center)
        self.condition_editor = ConditionEditor("단일 파일 조건")
        self.condition_editor.set_custom_values_getter(self.get_single_unique_values)
        center_l.addWidget(self.condition_editor)
        split.addWidget(center)

        right = make_card(); right_l = QVBoxLayout(right)
        right_l.addWidget(QLabel("정렬 설정"))
        row = QHBoxLayout()
        self.cmb_sort_col = QComboBox(); row.addWidget(self.cmb_sort_col)
        self.cmb_sort_order = QComboBox(); self.cmb_sort_order.addItems(["asc", "desc"]); row.addWidget(self.cmb_sort_order)
        self.btn_add_sort = QPushButton("정렬 추가"); self.btn_add_sort.clicked.connect(self.add_sort_row); row.addWidget(self.btn_add_sort)
        right_l.addLayout(row)

        self.tbl_sorts = QTableWidget(0, 3)
        self.tbl_sorts.setHorizontalHeaderLabels(["사용", "컬럼", "정렬"])
        self.tbl_sorts.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.tbl_sorts.verticalHeader().setVisible(False)
        self.tbl_sorts.setSelectionBehavior(QTableWidget.SelectRows)
        right_l.addWidget(self.tbl_sorts)

        self.btn_remove_sort = QPushButton("선택 정렬 삭제")
        self.btn_remove_sort.clicked.connect(self.remove_selected_sort_rows)
        right_l.addWidget(self.btn_remove_sort)

        dedup_row = QHBoxLayout()
        self.cmb_dedup_col = QComboBox(); dedup_row.addWidget(self.cmb_dedup_col)
        self.cmb_dedup_keep = QComboBox(); self.cmb_dedup_keep.addItems(["first", "last"]); dedup_row.addWidget(self.cmb_dedup_keep)
        self.chk_enable_dedup = QCheckBox("중복 제거 사용"); dedup_row.addWidget(self.chk_enable_dedup)
        right_l.addLayout(dedup_row)
        split.addWidget(right)

        split.setStretchFactor(0, 7)
        split.setStretchFactor(1, 3)
        split.setStretchFactor(2, 4)
        root.addWidget(split, 4)

        preview_card = make_card(); preview_l = QVBoxLayout(preview_card)
        preview_l.addWidget(QLabel("미리보기"))
        self.txt_preview = QTextEdit(); self.txt_preview.setReadOnly(True)
        preview_l.addWidget(self.txt_preview)
        root.addWidget(preview_card, 2)

    def log(self, msg):
        self.log_cb(f"[단일] {msg}")

    def get_header_row_index(self):
        return self.cmb_header.currentData()

    def select_file(self):
        file_path, _ = QFileDialog.getOpenFileName(
            self, "단일 파일 선택", "", "Data Files (*.csv *.xlsx *.xlsm *.xls *.html *.htm)"
        )
        if not file_path:
            return
        self.file_path = file_path
        suffix = " [열려 있음]" if is_file_locked(file_path) else ""
        self.lbl_file.setText(file_path + suffix)
        self.cmb_sheet.clear()
        readable = self.get_readable_source()
        if readable:
            names = get_sheet_names(readable)
            if names:
                self.cmb_sheet.addItems(names)
            else:
                self.cmb_sheet.addItem("(기본)")
        self.refresh_all()

    def get_readable_source(self):
        if not self.file_path:
            return None
        return get_readable_file_path(self.file_path)

    def get_selected_sheet_name(self):
        t = self.cmb_sheet.currentText()
        return None if t == "(기본)" else clean_text(t)

    def get_single_unique_values(self, column_name):
        readable = self.get_readable_source()
        if not readable:
            return []
        return extract_unique_values_fast(
            readable,
            column_name,
            sheet_name=self.get_selected_sheet_name(),
            max_values=200,
            max_scan_rows=50000,
            header_row_idx=self.get_header_row_index(),
            force_html=self.chk_force_html.isChecked()
        )

    def refresh_all(self):
        readable = self.get_readable_source()
        if not readable:
            return
        try:
            cols = extract_columns_fast(
                readable,
                sheet_name=self.get_selected_sheet_name(),
                header_row_idx=self.get_header_row_index(),
                force_html=self.chk_force_html.isChecked()
            )
            prev = set(self.col_grid.get_checked_columns())
            self.col_grid.set_columns(cols, checked=True, preserve_checked=prev)
            self.condition_editor.set_columns(cols)
            for combo in [self.cmb_sort_col, self.cmb_dedup_col]:
                combo.clear()
                combo.addItems(cols)

            df = load_file_to_df(
                readable,
                sheet_name=self.get_selected_sheet_name(),
                header_row_idx=self.get_header_row_index(),
                force_html=self.chk_force_html.isChecked()
            )
            self.txt_preview.setPlainText(preview_text(df))
            if not cols:
                self.log("컬럼 인식 실패 → 헤더 행 또는 HTML 강제 읽기 옵션을 확인하세요.")
        except Exception as e:
            self.txt_preview.setPlainText(f"미리보기 오류: {e}")
            self.log(f"미리보기 오류: {e}")

    def show_raw_structure(self):
        readable = self.get_readable_source()
        if not readable:
            return
        try:
            rows = load_file_sample_rows(
                readable,
                sheet_name=self.get_selected_sheet_name(),
                max_rows=20,
                force_html=self.chk_force_html.isChecked()
            )
            txt = raw_rows_preview_text(rows, max_rows=20)
            QMessageBox.information(self, "원본 구조 보기", txt)
        except Exception as e:
            QMessageBox.warning(self, "오류", str(e))

    def append_sort_row(self, use=True, column="", order="asc"):
        row = self.tbl_sorts.rowCount()
        self.tbl_sorts.insertRow(row)
        chk = QTableWidgetItem()
        chk.setFlags(chk.flags() | Qt.ItemIsUserCheckable | Qt.ItemIsEnabled)
        chk.setCheckState(Qt.Checked if use else Qt.Unchecked)
        self.tbl_sorts.setItem(row, 0, chk)
        self.tbl_sorts.setItem(row, 1, QTableWidgetItem(str(column)))
        self.tbl_sorts.setItem(row, 2, QTableWidgetItem(str(order)))

    def add_sort_row(self):
        col = clean_text(self.cmb_sort_col.currentText())
        order = clean_text(self.cmb_sort_order.currentText()) or "asc"
        if not col:
            QMessageBox.warning(self, "알림", "정렬 컬럼을 선택하세요.")
            return
        self.append_sort_row(True, col, order)

    def remove_selected_sort_rows(self):
        selected = self.tbl_sorts.selectionModel().selectedRows()
        for row in sorted([x.row() for x in selected], reverse=True):
            self.tbl_sorts.removeRow(row)

    def parse_sorts(self):
        result = []
        for row in range(self.tbl_sorts.rowCount()):
            use_item = self.tbl_sorts.item(row, 0)
            if not use_item or use_item.checkState() != Qt.Checked:
                continue
            col = clean_text(self.tbl_sorts.item(row, 1).text()) if self.tbl_sorts.item(row, 1) else ""
            order = clean_text(self.tbl_sorts.item(row, 2).text()).lower() if self.tbl_sorts.item(row, 2) else "asc"
            if col:
                result.append({"column": col, "order": "desc" if order == "desc" else "asc"})
        return result

    def parse_dedup(self):
        if not self.chk_enable_dedup.isChecked():
            return None
        col = clean_text(self.cmb_dedup_col.currentText())
        keep = clean_text(self.cmb_dedup_keep.currentText()).lower() or "first"
        if not col:
            return None
        return {"column": col, "keep": "last" if keep == "last" else "first"}

    def run_export(self):
        readable = self.get_readable_source()
        if not readable:
            QMessageBox.warning(self, "알림", "파일을 선택하세요.")
            return
        selected_columns = self.col_grid.get_checked_columns()
        if not selected_columns:
            QMessageBox.warning(self, "알림", "출력할 컬럼을 선택하세요.")
            return
        output_path, _ = QFileDialog.getSaveFileName(self, "저장 파일", "결과.csv", "CSV Files (*.csv)")
        if not output_path:
            return
        self.worker = ExportWorker(
            source_path=self.file_path,
            output_path=output_path,
            sheet_name=self.get_selected_sheet_name(),
            selected_columns=selected_columns,
            conditions=self.condition_editor.parse_conditions(),
            fill_service=self.chk_fill_service.isChecked(),
            sort_specs=self.parse_sorts(),
            dedup_spec=self.parse_dedup(),
            header_row_idx=self.get_header_row_index(),
            force_html=self.chk_force_html.isChecked(),
        )
        self.worker.status_changed.connect(self.log_cb)
        self.worker.finished_ok.connect(lambda msg: QMessageBox.information(self, "완료", msg))
        self.worker.error_occurred.connect(lambda msg: QMessageBox.critical(self, "오류", msg))
        self.worker.start()


class OpenExcelTab(QWidget):
    def __init__(self, log_cb):
        super().__init__()
        self.log_cb = log_cb
        self.open_items = []
        self.build_ui()

    def build_ui(self):
        root = QVBoxLayout(self)
        top = QHBoxLayout()
        self.btn_refresh = QPushButton("열려있는 엑셀 새로고침")
        self.btn_refresh.clicked.connect(self.refresh_open_workbooks)
        top.addWidget(self.btn_refresh)

        self.cmb_workbook = QComboBox()
        self.cmb_workbook.currentIndexChanged.connect(self.on_workbook_changed)
        top.addWidget(self.cmb_workbook)

        self.cmb_sheet = QComboBox()
        self.cmb_sheet.currentIndexChanged.connect(self.refresh_preview)
        top.addWidget(self.cmb_sheet)

        top.addWidget(QLabel("헤더"))
        self.cmb_header = QComboBox()
        self.cmb_header.addItem("자동", None)
        for i in range(1, 11):
            self.cmb_header.addItem(f"{i}행", i - 1)
        self.cmb_header.currentIndexChanged.connect(self.refresh_preview)
        top.addWidget(self.cmb_header)

        self.chk_fill_service = QCheckBox("서비스(소) 자동 채움")
        self.chk_fill_service.setChecked(True)
        top.addWidget(self.chk_fill_service)

        self.btn_diag = QPushButton("원본 구조 보기")
        self.btn_diag.clicked.connect(self.show_raw_structure)
        top.addWidget(self.btn_diag)

        top.addStretch()

        self.btn_extract = QPushButton("열린 엑셀 추출")
        self.btn_extract.clicked.connect(self.run_export)
        top.addWidget(self.btn_extract)
        root.addLayout(top)

        if not WIN32_AVAILABLE:
            warn = QLabel("pywin32가 설치되어 있지 않아 '열려있는 엑셀 인식' 기능은 비활성 상태입니다.\npip install pywin32 후 사용하세요.")
            root.addWidget(warn)

        split = QSplitter(Qt.Horizontal)
        left = make_card(); left_l = QVBoxLayout(left)
        self.col_grid = ColumnCheckGrid("열린 시트의 컬럼", max_columns=4)
        left_l.addWidget(self.col_grid)
        split.addWidget(left)

        center = make_card(); center_l = QVBoxLayout(center)
        self.condition_editor = ConditionEditor("열린 엑셀 조건")
        self.condition_editor.set_custom_values_getter(self.get_open_excel_unique_values)
        center_l.addWidget(self.condition_editor)
        split.addWidget(center)

        right = make_card(); right_l = QVBoxLayout(right)
        right_l.addWidget(QLabel("미리보기"))
        self.txt_preview = QTextEdit(); self.txt_preview.setReadOnly(True)
        right_l.addWidget(self.txt_preview)
        split.addWidget(right)

        split.setStretchFactor(0, 7)
        split.setStretchFactor(1, 3)
        split.setStretchFactor(2, 4)
        root.addWidget(split, 4)

        self.refresh_open_workbooks()

    def log(self, msg):
        self.log_cb(f"[열린엑셀] {msg}")

    def get_header_row_index(self):
        return self.cmb_header.currentData()

    def refresh_open_workbooks(self):
        self.open_items = list_open_excel_workbooks()
        self.cmb_workbook.clear()
        if not self.open_items:
            self.cmb_workbook.addItem("(열려있는 엑셀 없음)")
            self.cmb_sheet.clear()
            self.col_grid.set_columns([], checked=False)
            self.condition_editor.set_columns([])
            self.txt_preview.setPlainText("")
            return
        for item in self.open_items:
            label = item['name']
            if not item['full_name']:
                label += " [저장안됨]"
            self.cmb_workbook.addItem(label)
        self.on_workbook_changed()

    def get_current_item(self):
        idx = self.cmb_workbook.currentIndex()
        if idx < 0 or idx >= len(self.open_items):
            return None
        return self.open_items[idx]

    def on_workbook_changed(self):
        item = self.get_current_item()
        self.cmb_sheet.clear()
        if not item:
            return
        self.cmb_sheet.addItems(item.get("sheet_names", []))
        self.refresh_preview()

    def get_open_excel_unique_values(self, column_name):
        item = self.get_current_item()
        if not item:
            return []
        sheet_name = clean_text(self.cmb_sheet.currentText())
        if not sheet_name:
            return []
        return extract_open_excel_unique_values(
            item["name"],
            sheet_name,
            column_name,
            max_values=200,
            max_scan_rows=50000,
            header_row_idx=self.get_header_row_index()
        )

    def refresh_preview(self):
        item = self.get_current_item()
        if not item:
            self.txt_preview.setPlainText("")
            self.col_grid.set_columns([], checked=False)
            self.condition_editor.set_columns([])
            return

        sheet_name = clean_text(self.cmb_sheet.currentText())
        if not sheet_name:
            self.txt_preview.setPlainText("")
            return

        try:
            cols = extract_open_excel_columns(
                item["name"],
                sheet_name,
                scan_rows=200,
                header_row_idx=self.get_header_row_index()
            )
            prev = set(self.col_grid.get_checked_columns())
            self.col_grid.set_columns(cols, checked=True, preserve_checked=prev)
            self.condition_editor.value_cache.clear()
            self.condition_editor.set_columns(cols)

            df = load_open_excel_sheet_df(
                item["name"],
                sheet_name,
                max_rows=300,
                header_row_idx=self.get_header_row_index()
            )
            self.txt_preview.setPlainText(preview_text(df))

            if not cols:
                self.log("컬럼 인식 실패 → 헤더 행 선택 또는 원본 구조 보기를 확인하세요.")
        except Exception as e:
            self.txt_preview.setPlainText(f"미리보기 오류: {e}")
            self.log(f"미리보기 오류: {e}")

    def show_raw_structure(self):
        item = self.get_current_item()
        if not item:
            return
        sheet_name = clean_text(self.cmb_sheet.currentText())
        if not sheet_name:
            return
        try:
            rows = load_open_excel_raw_rows(item["name"], sheet_name, max_rows=20)
            txt = raw_rows_preview_text(rows, max_rows=20)
            QMessageBox.information(self, "원본 구조 보기", txt)
        except Exception as e:
            QMessageBox.warning(self, "오류", str(e))

    def run_export(self):
        item = self.get_current_item()
        if not item:
            QMessageBox.warning(self, "알림", "열려있는 엑셀을 찾지 못했습니다.")
            return
        sheet_name = clean_text(self.cmb_sheet.currentText())
        if not sheet_name:
            QMessageBox.warning(self, "알림", "시트를 선택하세요.")
            return
        selected_columns = self.col_grid.get_checked_columns()
        if not selected_columns:
            QMessageBox.warning(self, "알림", "출력할 컬럼을 선택하세요.")
            return
        output_path, _ = QFileDialog.getSaveFileName(self, "저장 파일", "열린엑셀_결과.csv", "CSV Files (*.csv)")
        if not output_path:
            return

        self.worker = OpenExcelExportWorker(
            workbook_name=item["name"],
            sheet_name=sheet_name,
            output_path=output_path,
            selected_columns=selected_columns,
            conditions=self.condition_editor.parse_conditions(),
            fill_service=self.chk_fill_service.isChecked(),
            header_row_idx=self.get_header_row_index(),
        )
        self.worker.status_changed.connect(self.log_cb)
        self.worker.finished_ok.connect(lambda msg: QMessageBox.information(self, "완료", msg))
        self.worker.error_occurred.connect(lambda msg: QMessageBox.critical(self, "오류", msg))
        self.worker.start()

def make_card():
    frame = QFrame()
    frame.setStyleSheet("""
        QFrame {
            border: 1px solid #d9e0ea;
            border-radius: 14px;
            background: #ffffff;
        }
    """)
    return frame


# =========================================================
# 메인 윈도우
# =========================================================
class MainWindow(QWidget):
    THEMES = {
        "Ocean": {
            "bg": "#f4f8fb", "card": "#ffffff", "line": "#dbe4ef", "text": "#1f2937", "primary": "#2563eb", "primary_hover": "#1d4ed8", "primary_press": "#1e40af", "chunk": "#22c55e", "header": "#eff6ff"
        },
        "Forest": {
            "bg": "#f5faf6", "card": "#ffffff", "line": "#d8e7db", "text": "#203126", "primary": "#2f855a", "primary_hover": "#276749", "primary_press": "#22543d", "chunk": "#38a169", "header": "#edf7f1"
        },
        "Sunset": {
            "bg": "#fff8f3", "card": "#ffffff", "line": "#f0ddd2", "text": "#3f2d22", "primary": "#dd6b20", "primary_hover": "#c05621", "primary_press": "#9c4221", "chunk": "#ed8936", "header": "#fff1e8"
        },
        "Plum": {
            "bg": "#faf7fc", "card": "#ffffff", "line": "#e7def0", "text": "#2d1f3a", "primary": "#805ad5", "primary_hover": "#6b46c1", "primary_press": "#553c9a", "chunk": "#9f7aea", "header": "#f3ecff"
        },
        "Dark Slate": {
            "bg": "#0f172a", "card": "#111827", "line": "#334155", "text": "#e5e7eb", "primary": "#3b82f6", "primary_hover": "#2563eb", "primary_press": "#1d4ed8", "chunk": "#22c55e", "header": "#1e293b"
        },
    }

    def __init__(self):
        super().__init__()
        self.setWindowTitle("통합 데이터 병합·추출기")
        self.resize(1760, 1080)
        self._pending_logs = []
        self.txt_log = None
        self.build_ui()
        self.apply_theme("Ocean")

    def build_ui(self):
        root = QVBoxLayout(self)

        top = QHBoxLayout()
        title = QLabel("통합 데이터 병합·추출기")
        top.addWidget(title)
        top.addStretch()
        top.addWidget(QLabel("테마"))
        self.cmb_theme = QComboBox()
        self.cmb_theme.addItems(list(self.THEMES.keys()))
        self.cmb_theme.currentTextChanged.connect(self.apply_theme)
        top.addWidget(self.cmb_theme)
        root.addLayout(top)

        self.progress = QProgressBar()
        self.progress.setRange(0, 100)
        root.addWidget(self.progress)

        log_card = make_card()
        log_layout = QVBoxLayout(log_card)
        log_layout.addWidget(QLabel("통합 로그"))
        self.txt_log = QTextEdit(); self.txt_log.setReadOnly(True)
        log_layout.addWidget(self.txt_log)

        self.tabs = QTabWidget()
        self.merge_tab = MergeTab(self.log)
        self.single_tab = SingleFileTab(self.log)
        self.open_excel_tab = OpenExcelTab(self.log)
        self.tabs.addTab(self.merge_tab, "1. 병합")
        self.tabs.addTab(self.single_tab, "2. 단일 파일 추출")
        self.tabs.addTab(self.open_excel_tab, "3. 열려있는 엑셀")
        root.addWidget(self.tabs, 5)
        root.addWidget(log_card, 2)

        if self._pending_logs:
            for _msg in self._pending_logs:
                self.txt_log.append(_msg)
            self._pending_logs.clear()

    def log(self, msg):
        if self.txt_log is None:
            self._pending_logs.append(msg)
            return
        self.txt_log.append(msg)

    def on_progress(self, value):
        self.progress.setValue(value)

    
    def apply_theme(self, name):
        t = self.THEMES.get(name, self.THEMES["Ocean"])
        self.setStyleSheet(f"""
            QWidget {{
                background-color: {t['bg']};
                color: {t['text']};
                font-family: 'Malgun Gothic';
                font-size: 10pt;
            }}
            QFrame {{
                background: {t['card']};
                border: 1px solid {t['line']};
                border-radius: 14px;
            }}
            QFrame#helpCard {{
                background: {t['header']};
                border: 1px solid {t['line']};
                border-radius: 12px;
            }}
            QLabel {{
                background: transparent;
                border: none;
                font-weight: 500;
            }}
            QLabel#helpTitle {{
                font-weight: 700;
                font-size: 10.5pt;
            }}
            QLabel#sectionTitle {{
                font-weight: 700;
                font-size: 11pt;
            }}
            QPushButton {{
                background-color: {t['primary']};
                color: white;
                border: none;
                border-radius: 10px;
                padding: 10px 14px;
                font-weight: 600;
                min-height: 18px;
            }}
            QPushButton:hover {{ background-color: {t['primary_hover']}; }}
            QPushButton:pressed {{ background-color: {t['primary_press']}; }}
            QLineEdit, QTextEdit, QListWidget, QComboBox, QTableWidget, QTabWidget::pane {{
                background: {t['card']};
                border: 1px solid {t['line']};
                border-radius: 10px;
                padding: 6px;
                selection-background-color: #dbeafe;
            }}
            QHeaderView::section {{
                background-color: {t['header']};
                border: 1px solid {t['line']};
                padding: 6px;
                font-weight: 600;
            }}
            QProgressBar {{
                background: {t['line']};
                border: none;
                border-radius: 8px;
                text-align: center;
                height: 18px;
                color: {t['text']};
                font-weight: 600;
            }}
            QProgressBar::chunk {{
                background-color: {t['chunk']};
                border-radius: 8px;
            }}
            QScrollArea {{ border: none; background: transparent; }}
            QCheckBox {{ spacing: 8px; background: transparent; }}
            QTabWidget::pane {{
                border: 1px solid {t['line']};
                border-radius: 12px;
                margin-top: 8px;
                background: {t['card']};
            }}
            QTabBar::tab {{
                background: {t['header']};
                border: 1px solid {t['line']};
                border-bottom: none;
                padding: 10px 18px;
                margin-right: 4px;
                border-top-left-radius: 10px;
                border-top-right-radius: 10px;
                font-weight: 600;
            }}
            QTabBar::tab:selected {{
                background: {t['card']};
                color: {t['text']};
            }}
            QTabBar::tab:!selected {{
                color: {t['text']};
            }}
        """)


APP_PASSWORD = "0303"


def request_startup_password(parent=None):
    password, ok = QInputDialog.getText(
        parent,
        "접속 확인",
        "접속 비밀번호를 입력하세요",
        QLineEdit.Password,
    )
    if not ok:
        return False
    return password == APP_PASSWORD


if __name__ == "__main__":

    app = QApplication(sys.argv)
    app.setFont(QFont("맑은 고딕", 10))

    if not request_startup_password():
        QMessageBox.warning(None, "접속 실패", "비밀번호가 올바르지 않거나 입력이 취소되었습니다.")
        sys.exit(0)

    win = MainWindow()
    win.show()
    sys.exit(app.exec())
